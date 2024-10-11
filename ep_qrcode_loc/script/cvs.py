import pandas as pd
import numpy as np
import openpyxl
import csv

def csv2xlsx(txt_file, excel_file):
    df = pd.read_csv(txt_file, sep = ',')
    df.to_excel(excel_file, index = False)
    work_book = openpyxl.load_workbook(excel_file)
    sheet = work_book.active
    #先将现有数据下移一行
    max_row = sheet.max_row
    max_column = sheet.max_column   
    for row in range(max_row,  0, -1):
        for col in range(1, max_column + 1):
            current_cell_value = sheet.cell(row = row, column = col).value
            sheet.cell(row = row + 1, column = col, value = current_cell_value)
    for row in range(1, max_row + 1):
        for col in range(1, max_column + 1):
            sheet.cell(row = row, column = col, value = None)
            
    #将第一行填充为行名
    column_names = ["time", "index", "vel_x", "des_err", "camera_x_err", "camera_y_err"]
    for col_num, col_name in enumerate(column_names, 1):
        sheet.cell(row = 1, column = col_num, value = col_name)
    work_book.save(excel_file)
    print("处理完成")
    
def IsFloatNum(str):
    s = str.split('.')
    # print("snum = ", s)
    if len(s) >= 2:
        return True
    else:
        for si in s:
            if not si.isdigit():
                return False
        return False
 
def process_(excel_file,final_file,target_file):
    # 读取Excel文件  
    work_book = openpyxl.load_workbook(excel_file)
    sheet = work_book.active
    max_row = sheet.max_row
    max_column = sheet.max_column 
    df1 = pd.read_excel(excel_file) 
    indexes = df1['index'].to_numpy() 
    vel_x = df1['vel_x'].to_numpy()
    des_err = df1['des_err'].to_numpy()
    mean_err = np.zeros((2700, 1))
    err_count = np.zeros((2700, 1))
    qrcode_len = 0
    for row in range(0, max_row-1):
        idx = indexes[row] - 100000
        # if(idx == 782): print(f'idx = {idx}')
        if vel_x[row] >= 0:
            mean_err[idx] += des_err[row]
            err_count[idx] += 1
    
    for i in range(err_count.shape[0]):
        if(err_count[i] > 0):
            # if(i == 782): print(f'mean_err = {mean_err[i]}\terr_count = {err_count[i]}')
            mean_err[i] /= err_count[i]
        
    df = pd.read_csv(target_file, sep = ',', dtype = object)
    print(df.iloc[:,3].shape[0])
    print(df.iat[69,3])
    s = df.iat[69,3].split('.')
    print(len(s))
    if(IsFloatNum(df.iloc[69,3]) == False): print("yes")
    for i in range(df.iloc[:,3].shape[0]):
        if(pd.isnull(df.iloc[i,3]) == False and IsFloatNum(df.iloc[i,3]) == False and int(df.iloc[i,3]) >= 100000):
            idx = int(int(df.iloc[i,3]) - 100000)
            df.iat[i,6] = str(np.round(mean_err[idx,0],6))
    df.to_csv('NewSiteTable2.csv', index = False, header = 0, sep = ' ')

if __name__ == "__main__":
    txt_file = "/home/hzz/zlProjects/read_excel/err2.csv"
    excel_file = "/home/hzz/zlProjects/read_excel/demo_output2.xlsx"
    final_file = "/home/hzz/zlProjects/read_excel/final_2csv"
    target_file = "/home/hzz/zlProjects/read_excel/SiteTable.csv"
    csv2xlsx(txt_file, excel_file)
    process_(excel_file,final_file,target_file)
    
    # 102227
    # 最大是102699